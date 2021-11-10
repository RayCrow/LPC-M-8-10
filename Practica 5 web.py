import requests
import bs4 as bs
from openpyxl import Workbook


drama = []
drama2 = []
comedia = []
comedia2 = []
animacion = []
animacion2 = []
kidyfam = []
kidyfam2 = []
lenguaext = []
lenguaext2 = []
reality = []
reality2 = []
continuacion = []
continuacion2 = []
endesarrollo = []
endesarrollo2 = []


# GENERADOR DE EXCEL / TABLA DEL GENERO DE PELICULA-SERIE ELEGIDO
def excel(lista,name):
    libro = Workbook()
    pagina = libro.active
    x = 1
    pagina.cell(row=1, column=2, value=name)
    for elem in lista:
        for i in range(0,len(elem)):
            pagina.cell(row=x+1, column=i+1, value=elem[i])
        x = x + 1
    libro.save("NetflixMovies.xlsx")
    print("Excel generado!")


url = "https://es.wikipedia.org/wiki/Anexo:Programaci%C3%B3n_original_distribuida_por_Netflix"
page = requests.get(url)

if page.status_code == 200:
    soup = bs.BeautifulSoup(page.content, 'html.parser')

    print("PROXIMOS ESTRENOS DE NETFLIX. Elige el género")
    var = int(input("""
                    1) Drama
                    2) Comedia
                    3) Animación
                    4) Niños y Familia
                    5) Lengua extranjera
                    6) RealityShow
                    7) Continuaciónes
                    8) En desarrollo
                    Opcion: """))
    if var == 1:
        # DRAMA
        name = " PROXMIOS ESTRENOS: DRAMA"
        tab1 = soup.select("table")[27].find_all("tr")
        for i in tab1:
            drama.append(i.getText().split("\n"))
        for i in drama:
            i = list(i)
            drama2.append(i)
        excel(drama2,name)
    elif var == 2:
        # COMEDIA
        name = " PROXMIOS ESTRENOS: COMEDIA"
        tab2 = soup.select("table")[28].find_all("tr")
        for i in tab2:
            comedia.append(i.getText().split("\n"))
        for i in comedia:
            i = list(i)
            comedia2.append(i)
        excel(comedia2,name)
    elif var == 3:
        # ANIMACION
        name = " PROXMIOS ESTRENOS: ANIMACION"
        tab3 = soup.select("table")[29].find_all("tr")
        for i in tab3:
            animacion.append(i.getText().split("\n"))
        for i in animacion:
            i = list(i)
            animacion2.append(i)
        excel(animacion2,name)
    elif var == 4:
        # NIÑOS Y FAMILIA
        name = " PROXMIOS ESTRENOS: NIÑOS Y FAMILIA"
        tab4 = soup.select("table")[31].find_all("tr")
        for i in tab4:
            kidyfam.append(i.getText().split("\n"))
        for i in kidyfam:
            i = list(i)
            kidyfam2.append(i)
        excel(kidyfam2,name)
    elif var == 5:
        # LENGUA EXTRANJERA
        name = " PROXMIOS ESTRENOS: LENGUA EXTRANJERA"
        tab5 = soup.select("table")[32].find_all("tr")
        for i in tab5:
            lenguaext.append(i.getText().split("\n"))
        for i in lenguaext:
            i = list(i)
            lenguaext2.append(i)
        excel(lenguaext2,name)
    elif var == 6:
        # REALITYSHOW
        name = " PROXMIOS ESTRENOS: REALITY-SHOW"
        tab6 = soup.select("table")[34].find_all("tr")
        for i in tab6:
            reality.append(i.getText().split("\n"))
        for i in reality:
            i = list(i)
            reality2.append(i)
        excel(reality2,name)
    elif var == 7:
        # CONTINUACIONES
        name = " PROXMIOS ESTRENOS: CONTINUACIONES"
        tab7 = soup.select("table")[37].find_all("tr")
        for i in tab7:
            continuacion.append(i.getText().split("\n"))
        for i in continuacion:
            i = list(i)
            continuacion2.append(i)
        excel(continuacion2,name)
    elif var == 8:
        # EN DESARROLLO
        name = " PROXMIOS ESTRENOS: EN DESARROLLO"
        tab5 = soup.select("table")[38].find_all("tr")
        for i in tab5:
            endesarrollo.append(i.getText().split("\n"))
        for i in endesarrollo:
            i = list(i)
            endesarrollo2.append(i)
        excel(endesarrollo2,name)
    else:
        print("Opción no valida. Adiós!")
        exit()
else:
    exit()